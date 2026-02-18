import os
from openpyxl import load_workbook, Workbook
from datetime import datetime

EXCEL_PATH = "data/master.xlsx"

HEADERS = [
    "Name",
    "Company",
    "Title",
    "Email",
    "Phone",
    "CreatedAt",
    "UpdatedAt",
    "Source"
]

def ensure_excel_exists():
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)

def upsert_contact(data: dict, source: str):

    ensure_excel_exists()

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    email = data.get("Email")
    phone = data.get("Phone")

    found_row = None

    for row in ws.iter_rows(min_row=2):
        existing_email = row[3].value
        existing_phone = row[4].value

        if email and existing_email == email:
            found_row = row
            break
        if phone and existing_phone == phone:
            found_row = row
            break

    now = datetime.utcnow().isoformat()

    if found_row:
        if not found_row[0].value:
            found_row[0].value = data.get("Name")
        if not found_row[1].value:
            found_row[1].value = data.get("Company")
        if not found_row[2].value:
            found_row[2].value = data.get("Title")

        found_row[6].value = now
        status = "updated"

    else:
        ws.append([
            data.get("Name"),
            data.get("Company"),
            data.get("Title"),
            data.get("Email"),
            data.get("Phone"),
            now,
            now,
            source
        ])
        status = "inserted"

    wb.save(EXCEL_PATH)
    return status
